import os
import csv
from typing import Dict, List, Iterable
from fastapi import HTTPException, status
from openpyxl import load_workbook


def _canonicalize_key(key: str) -> str:
    return key.strip().lower().replace(" ", "_")


def _normalize_record(record: Dict[str, str]) -> Dict[str, str]:
    canonical = { _canonicalize_key(k): (v if v is not None else "") for k, v in record.items() }

    key_map = {
        "start_location": ["start", "start_location", "startlocation", "from", "from_location"],
        "end_location": ["end", "end_location", "endlocation", "to", "to_location"],
        "price": ["price", "amount", "fare", "cost"],
    }

    normalized: Dict[str, str] = {"start_location": "", "end_location": "", "price": ""}
    for target, aliases in key_map.items():
        for alias in aliases:
            if alias in canonical and str(canonical[alias]).strip() != "":
                normalized[target] = str(canonical[alias])
                break

    missing = [k for k, v in normalized.items() if v == ""]
    if missing:
        raise ValueError(f"Missing or empty values for required columns: {', '.join(missing)}")

    return normalized


def _read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV file must have a header row.")
        records = [_normalize_record(row) for row in reader if any((v or "").strip() for v in row.values())]
        return records


def _read_excel(path: str) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows: Iterable = ws.iter_rows(values_only=True)
    try:
        header_keys = [str(h) if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []

    if not any(header_keys):
        raise ValueError("Excel sheet is empty or has no header row.")

    records: List[Dict[str, str]] = []
    for row in rows:
        if not any(str(v).strip() for v in (row or [])):
            continue
        record = {header_keys[i]: val for i, val in enumerate(row)}
        records.append(_normalize_record(record))
    return records


def load_and_validate_records(file_path: str) -> Dict:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Data file not found: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    records: List[Dict[str, str]] = []

    try:
        if ext == ".csv":
            records = _read_csv(file_path)
        elif ext in (".xlsx", ".xlsm", ".xls"):
            records = _read_excel(file_path)
        else:
            raise ValueError("Unsupported file type. Please provide a .csv or .xlsx file.")

        if not records:
            raise ValueError("No valid data rows found in the file.")

        return {
            "records": records,
            "columns": list(records[0].keys()),
            "total_rows": len(records),
            "preview": records[:3],
        }
    except (ValueError, KeyError, IndexError) as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"An unexpected error occurred: {e}")
